import openpyxl
book =openpyxl.load_workbook(r"C:\Users\amznkris\Documents\excelautomation\Demo.xlsx")
sheet =book.active
cell =sheet.cell(row=1, column=2)
cell2 =sheet.cell(row=1, column=3)
print(cell.value)
print(cell2.value)
sheet.cell(row=2, column=2).value = "Krishna"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_column)
print(sheet.max_row)
                       
for i in range(1,sheet.max_row+1):
    if sheet.cell(row =i, column=1).value == "case 6":  #to get the specifed column 
       for j in range(1,sheet.max_column+1):
           print(sheet.cell(row=i, column=j).value)
